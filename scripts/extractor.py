import os
import re
import tempfile
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, asdict

from fastapi import FastAPI, UploadFile, File, HTTPException
from PIL import Image

from docx import Document
from zipfile import ZipFile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

import pdfplumber
import pytesseract
from pdf2image import convert_from_path

"""
Universal form extractor service.
Start: uvicorn extractor:app --host 0.0.0.0 --port 8000
Requires tesseract + poppler installed on PATH.
"""


@dataclass
class ExtractedQuestion:
    text: str
    response_type: str
    options: List[str]
    confidence: float
    source: Dict[str, Any]
    debug: Dict[str, Any]


@dataclass
class ExtractedSubcategory:
    name: str
    questions: List[ExtractedQuestion]


@dataclass
class ExtractedCategory:
    name: str
    subcategories: List[ExtractedSubcategory]


app = FastAPI(title="Universal Form Extractor", version="1.0")


def _normalize(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[ \t]+", " ", s)
    return s


def clean_heading(t: str) -> str:
    t = _normalize(t)
    t = re.sub(r"SELF[- ]ASSESSMENT QUESTIONNAIRE", "", t, flags=re.I).strip()
    return t


def extract_text_from_docx(path: str) -> List[Dict[str, Any]]:
    doc = Document(path)
    out = []
    for i, p in enumerate(doc.paragraphs):
        t = _normalize(p.text)
        if t:
            out.append({"text": t, "source": {"type": "docx", "paragraph": i}})
    return out


def extract_text_from_xlsx(path: str, max_rows_per_sheet: int = 2000) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        row_count = 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_count >= max_rows_per_sheet:
                break
            cells = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
            if not cells:
                continue
            line = _normalize(" | ".join(cells))
            out.append({"text": line, "source": {"type": "xlsx", "sheet": sheet, "row": r_idx}})
            row_count += 1
    return out


def extract_text_from_pdf(path: str, ocr_if_needed: bool = True, ocr_page_limit: int = 60) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    text_chars = 0
    with pdfplumber.open(path) as pdf:
        page_count = len(pdf.pages)
        for pageno, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            txt = _normalize(txt)
            if txt:
                out.append({"text": txt, "source": {"type": "pdf", "page": pageno, "mode": "text"}})
                text_chars += len(txt)
    if ocr_if_needed and text_chars < 800:
        images = convert_from_path(path, dpi=250, first_page=1, last_page=min(ocr_page_limit, page_count))
        for idx, img in enumerate(images, start=1):
            ocr_txt = pytesseract.image_to_string(img)
            ocr_txt = _normalize(ocr_txt)
            if ocr_txt:
                out.append({"text": ocr_txt, "source": {"type": "pdf", "page": idx, "mode": "ocr"}})
    return out


def extract_text_from_image(path: str) -> List[Dict[str, Any]]:
    img = Image.open(path)
    txt = pytesseract.image_to_string(img)
    txt = _normalize(txt)
    if not txt:
        return []
    return [{"text": txt, "source": {"type": "image", "mode": "ocr"}}]


def extract_text_universal(path: str) -> List[Dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return extract_text_from_docx(path)
    if ext in [".xlsx", ".xlsm"]:
        return extract_text_from_xlsx(path)
    if ext == ".pdf":
        return extract_text_from_pdf(path, ocr_if_needed=True)
    if ext in [".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"]:
        return extract_text_from_image(path)
    raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")


Q_LINE_PAT = re.compile(
    r"""^(
        (\d{1,3}\s*[\.\)]\s+.+\?)|
        (\d{1,3}\s+.+\?)|
        (.+\?)|
        (Does|Do|Is|Are|Has|Have|Can|Please)\b.+
    )$""",
    re.I | re.X,
)
YES_NO_NA_PAT = re.compile(r"\bYes\b\s+\bNo\b\s+\bNA\b", re.I)
YES_NO_PAT = re.compile(r"\bYes\b\s+\bNo\b", re.I)
SECTION_LIKE_PAT = re.compile(r"^[A-Z][A-Z 0-9&/,:\\-]{6,140}$")
TITLE_LIKE_PAT = re.compile(r"^[A-Z][a-z].{2,140}$")
BULLET_PAT = re.compile(r"^(\-|\•|\*|\u25a1|\u2610|\u2611)\s+")
CHOICE_INLINE_PAT = re.compile(r"(Yes\s+No\s+NA|Yes\s+No)", re.I)


def infer_response_type_and_options(context_lines: List[str]) -> (str, List[str], float):
    joined = " ".join(context_lines)
    if YES_NO_NA_PAT.search(joined):
        return "yes_no_na", ["Yes", "No", "NA"], 0.90
    if YES_NO_PAT.search(joined):
        yn_count = len(re.findall(r"\bYes\b\s+\bNo\b", joined, flags=re.I))
        if yn_count <= 2:
            return "yes_no", ["Yes", "No"], 0.85
        return "mixed", ["Yes", "No"], 0.80
    opts = []
    for ln in context_lines:
        t = _normalize(ln)
        if not t:
            continue
        if BULLET_PAT.match(t):
            t = _normalize(BULLET_PAT.sub("", t))
        if 2 <= len(t) <= 120 and not t.endswith("?"):
            if re.match(r"^(Please\s+explain|Comments?|Please\s+indicate)\b", t, flags=re.I):
                continue
            opts.append(t)
    seen = set()
    uniq = []
    for o in opts:
        key = o.lower()
        if key not in seen:
            uniq.append(o)
            seen.add(key)
    if len(uniq) >= 4:
        return "multi_select", uniq, 0.70
    if 2 <= len(uniq) <= 3:
        return "single_select", uniq, 0.60
    return "text", [], 0.55


def extract_structure_generic(text_items: List[Dict[str, Any]]) -> List[ExtractedCategory]:
    categories: List[ExtractedCategory] = []
    current_cat: Optional[ExtractedCategory] = None
    current_sub: Optional[ExtractedSubcategory] = None
    lines = [(ti["text"], ti.get("source", {})) for ti in text_items]

    def ensure_cat(name: str):
        nonlocal current_cat, current_sub
        if current_cat is None or current_cat.name != name:
            current_cat = ExtractedCategory(name=name, subcategories=[])
            categories.append(current_cat)
            current_sub = None

    def ensure_sub(name: str):
        nonlocal current_cat, current_sub
        if current_cat is None:
            ensure_cat("General")
        if current_sub is None or current_sub.name != name:
            current_sub = ExtractedSubcategory(name=name, questions=[])
            current_cat.subcategories.append(current_sub)

    ensure_cat("General")
    ensure_sub("General")

    i = 0
    while i < len(lines):
        line, src = lines[i]
        t = _normalize(line)
        if not t:
            i += 1
            continue

        # Category detection
        if SECTION_LIKE_PAT.match(t) and len(t) <= 140:
            heading = clean_heading(t)
            if heading:
                ensure_cat(heading.title())
                ensure_sub("General")
                i += 1
                continue

        # Subcategory detection (allow ALLCAPS or Title case)
        if (TITLE_LIKE_PAT.match(t) or SECTION_LIKE_PAT.match(t)) and not t.endswith("?") and len(t) <= 140:
            if not re.match(r"^(Comments?|Auditor|Verification|Guidance|Yes/No)\b", t, re.I):
                heading = clean_heading(t)
                if heading:
                    ensure_sub(heading)
                    i += 1
                    continue

        # Question detection
        q_candidate = t
        if Q_LINE_PAT.match(q_candidate) and ("?" in q_candidate or re.match(r"^(Does|Do|Is|Are|Has|Have|Can|Please)\b", q_candidate, re.I)):
            context: List[str] = []
            for j in range(i + 1, min(i + 10, len(lines))):
                nxt, _ = lines[j]
                nxt_norm = _normalize(nxt)
                if not nxt_norm:
                    continue
                if SECTION_LIKE_PAT.match(nxt_norm) or (TITLE_LIKE_PAT.match(nxt_norm) and not nxt_norm.endswith("?") and len(nxt_norm) <= 140):
                    break
                if Q_LINE_PAT.match(nxt_norm) and ("?" in nxt_norm):
                    break
                context.append(nxt_norm)
            if CHOICE_INLINE_PAT.search(q_candidate):
                context = [q_candidate] + context
            rtype, opts, conf = infer_response_type_and_options(context)
            ensure_sub(current_sub.name if current_sub else "General")
            current_sub.questions.append(
                ExtractedQuestion(
                    text=q_candidate,
                    response_type=rtype,
                    options=opts,
                    confidence=conf,
                    source=src,
                    debug={"context_used": context[:12]},
                )
            )
            i += 1
            continue
        i += 1

    cleaned: List[ExtractedCategory] = []
    for c in categories:
        scs = [sc for sc in c.subcategories if sc.questions]
        if scs:
            cleaned.append(ExtractedCategory(name=c.name, subcategories=scs))
    return cleaned


# ----------------------------
# PSCI DOCX parser (XML-based)
# ----------------------------

SECTION_MARKERS = [
    ("A", "MANAGEMENT SYSTEMS Self-Assessment Questionnaire", "Management Systems"),
    ("B", "ETHICS Self-Assessment Questionnaire", "Ethics"),
    ("C", "LABOR Self-Assessment Questionnaire", "Labor"),
    ("D", "ENVIRONMENTAL PROTECTION Self-Assessment Questionnaire", "Environmental Protection"),
    ("E", "HEALTH & SAFETY COMPLIANCE AND RISK MANAGEMENT Self-Assessment Questionnaire", "Health & Safety Compliance and Risk Management"),
    ("F", "Biological Safety", "Biological Safety"),
]

IGNORE_SUBSEC = {
    "Auditor Verification",
    "Please provide observations, details, comments and any supporting documents",
    "AUDITOR GUIDANCE",
    "AUDITOR GUDIANCE",
}

AUDITOR_COL_MARKERS = {
    "Auditor Verification",
    "Please provide observations, details, comments and any supporting documents",
}


def clean(txt: str) -> str:
    return re.sub(r"\s+", " ", (txt or "").strip())


def get_text(el) -> str:
    texts = []
    if el.tag.endswith('t'):
        texts.append(el.text or '')
    for child in el:
        texts.append(get_text(child))
    return ''.join(texts)


def iter_block_items(doc_tree):
    body = doc_tree.find('w:body', {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"})
    for child in body:
        tag = child.tag
        if tag.endswith('p'):
            yield {'type': 'p', 'element': child}
        elif tag.endswith('tbl'):
            yield {'type': 'tbl', 'element': child}


def paragraph_text(p_el) -> str:
    return clean(get_text(p_el))


def table_to_rows(tbl_el):
    rows = []
    for tr in tbl_el.findall('w:tr', {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
        row_cells = []
        for tc in tr.findall('w:tc', {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
            cell_text = clean(get_text(tc))
            row_cells.append(cell_text)
        rows.append(row_cells)
    return rows


def is_question_start(qnum_str: str, cells) -> bool:
    if not re.match(r"^\d{1,3}$", qnum_str):
        return False
    c1 = cells[1] if len(cells) > 1 else ''
    if not c1:
        return False
    if '?' in c1:
        return True
    return bool(re.match(r"^(Does|Do|Has|Have|Is|Are|Indicate|Please|Which|What|When|Where|How|Describe)\b", c1))


def summarize_response(lines):
    t = ' '.join(lines)
    parts = set()
    if re.search(r"\bNA\b", t):
        parts.add("Yes/No/NA")
    elif re.search(r"Yes\s+No", t):
        parts.add("Yes/No")
    if re.search(r"Please explain|Please describe|Comments", t, re.I):
        parts.add("explanation/comments text")
    if re.search(r"Frequency|Regular\s+Annual|annually|every", t, re.I):
        parts.add("frequency field")
    if re.search(r"list|please list|provide.*copy|web link|attach", t, re.I):
        parts.add("list / link / attachment field")
    if re.search(r"check all that apply|By on-site|By desk-top", t, re.I):
        parts.add("checklist items")
    if re.search(r"How many|number of|Specify|Total|Age", t, re.I):
        parts.add("numeric/text fields")
    return ', '.join(sorted(parts)) if parts else 'text / structured fields'


def infer_type_from_summary(summary: str):
    s = summary.lower()
    if "yes/no/na" in s:
        return "yes_no_na"
    if "yes/no" in s:
        return "yes_no"
    if "checklist" in s:
        return "multi_select"
    return "text"


def extract_questions_from_table(rows, section_code, section_name):
    questions = []
    current_subsec = None
    auditor_col_idxs = set()
    for row in rows:
        for i, v in enumerate(row):
            if v in AUDITOR_COL_MARKERS:
                auditor_col_idxs.add(i)
        if auditor_col_idxs:
            break

    def strip_auditor_cells(cell_row):
        return [v for idx, v in enumerate(cell_row) if idx not in auditor_col_idxs and v not in AUDITOR_COL_MARKERS]

    for row in rows:
        cells = strip_auditor_cells(row)
        if not cells:
            continue
        first = cells[0]
        if first and not re.match(r"^\d{1,3}$", first):
            uniq = {c for c in cells if c}
            if (
                len(uniq) == 1
                and first not in IGNORE_SUBSEC
                and 'Self-Assessment Questionnaire' not in first
                and 'AUDITOR' not in first
                and not re.search(r'Yes\s*No', first)
            ):
                current_subsec = first
            continue
        if is_question_start(first, cells):
            qtext = cells[1].split('AUDITOR GUIDANCE')[0].strip()
            option_lines = []
            for c in cells[1:]:
                if not c or c == qtext:
                    continue
                if 'AUDITOR GUIDANCE' in c or 'AUDITOR GUDIANCE' in c:
                    continue
                option_lines.append(c)
            summary = summarize_response(option_lines)
            response_type = infer_type_from_summary(summary)
            questions.append({
                'section': section_code,
                'section_name': section_name,
                'subsection': current_subsec,
                'question': qtext,
                'response_type': response_type,
                'response_options_summary': summary,
            })
    dedup = {}
    for q in questions:
        key = (q['section'], q['question'])
        if key not in dedup:
            dedup[key] = q
    return list(dedup.values())


def build_structure_psci(docx_path: str) -> List[ExtractedCategory]:
    with ZipFile(docx_path) as z:
        with z.open('word/document.xml') as f:
            xml_data = f.read()
    doc_tree = ET.fromstring(xml_data)
    blocks = list(iter_block_items(doc_tree))
    starts = []
    for code, marker, name in SECTION_MARKERS:
        for i, b in enumerate(blocks):
            if b['type'] == 'p':
                text = paragraph_text(b['element'])
            else:
                text = '\n'.join([' | '.join(row) for row in table_to_rows(b['element'])])
            if marker in text:
                starts.append((code, name, i))
                break
    starts.sort(key=lambda x: x[2])
    all_q = []
    for idx, (code, name, start) in enumerate(starts):
        end = starts[idx+1][2] if idx+1 < len(starts) else len(blocks)
        section_slice = blocks[start:end]
        for b in section_slice:
            if b['type'] == 'tbl':
                rows = table_to_rows(b['element'])
                all_q.extend(extract_questions_from_table(rows, code, name))
    by_section = {}
    for q in all_q:
        by_section.setdefault(q['section_name'], []).append(q)
    categories = []
    for sec_name, qs in by_section.items():
        sub_map = {}
        for q in qs:
            sub = q['subsection'] or 'General'
            sub_map.setdefault(sub, []).append(q)
        cat = ExtractedCategory(name=sec_name, subcategories=[])
        for sub_name, items in sub_map.items():
            subcat = ExtractedSubcategory(name=sub_name, questions=[])
            for item in items:
                subcat.questions.append(
                    ExtractedQuestion(
                        text=item['question'],
                        response_type=item['response_type'],
                        options=[],
                        confidence=0.6,
                        source={},
                        debug={"response_summary": item['response_options_summary']},
                    )
                )
            cat.subcategories.append(subcat)
        categories.append(cat)
    return categories


@app.get("/health")
def health() -> Dict[str, bool]:
    return {"ok": True}


@app.post("/extract")
async def extract(file: UploadFile = File(...)) -> Dict[str, Any]:
    suffix = os.path.splitext(file.filename)[1].lower()
    if suffix not in [".pdf", ".docx", ".xlsx", ".xlsm", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"]:
        raise HTTPException(status_code=400, detail=f"Unsupported file: {file.filename}")

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        contents = await file.read()
        tmp.write(contents)
        tmp_path = tmp.name

    try:
        if suffix == ".docx":
            cats = build_structure_psci(tmp_path)
            return {
                "filename": file.filename,
                "category_count": len(cats),
                "categories": [asdict(c) for c in cats],
                "note": "PSCI DOCX parser used",
            }
        else:
            text_items = extract_text_universal(tmp_path)
            if not text_items:
                return {
                    "filename": file.filename,
                    "extracted_text_items": 0,
                    "categories": [],
                    "note": "No extractable text found (OCR may have failed).",
                }
            cats = extract_structure_generic(text_items)
            return {
              "filename": file.filename,
              "extracted_text_items": len(text_items),
              "category_count": len(cats),
              "categories": [asdict(c) for c in cats],
            }
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
